#!/usr/bin/env python3
"""
extraer_productos_excel.py
──────────────────────────────────────────────────────────────────────────────
Extrae los productos del archivo INVENTARIO 2026 1.xlsx y genera un JSON
limpio y estructurado que el seeder PHP de Laravel puede consumir.

Uso:
    .venv/bin/python3 scripts/extraer_productos_excel.py
    .venv/bin/python3 scripts/extraer_productos_excel.py --excel "otro_archivo.xlsx"
    .venv/bin/python3 scripts/extraer_productos_excel.py --output "otro_directorio/"

Dependencias:
    pip install openpyxl
    (ya está disponible en el .venv del proyecto)
──────────────────────────────────────────────────────────────────────────────
"""

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no está instalado.")
    print("Instálalo con: .venv/bin/pip install openpyxl")
    sys.exit(1)


# ─── Configuración de hojas del Excel ────────────────────────────────────────
# Formato: (nombre_marca_en_sistema, nombre_hoja_en_excel)
MARCA_SHEETS = [
    ("LG",             "LG"),
    ("SAMSUNG",        "SAMSUNG"),
    ("HAIER",          "HAIER"),
    ("DAEWOO",         "DAEWOO"),
    ("SONY-PANASONIC", "SONY-PANASONIC"),
    ("IRT-PHILIPS",    "IRT-PHILIPS"),
    ("TOSHIBA",        "TOSHIBA "),
    ("HISENSE",        "HISENSE "),
    ("SKYWORTH",       "SKYWORTH"),
    ("PREMIER",        "PREMIER"),
    ("MASTER G",       "MASTER G"),
    ("CHINO",          "CHINO"),
    ("AOC-JVC",        "AOC-JVC"),
]

# Valores por defecto asignados durante la extracción
STOCK_MINIMO_DEFAULT        = 2
STOCK_IDEAL_DEFAULT         = 10
TIEMPO_REPOSICION_DEFAULT   = 7
STOCKS_NEGATIVOS_A_CERO     = False  # False = conservar el valor negativo real del Excel


def safe_num(value, default=None):
    """Convierte un valor a float de forma segura, retorna default si falla."""
    if value is None:
        return default
    try:
        s = str(value).strip().replace(",", ".")
        f = float(s)
        return default if f != f else f  # NaN check
    except (ValueError, TypeError):
        return default


def parse_pulgadas(desc: str):
    """
    Extrae las pulgadas desde la descripción.
    Soporta múltiples formatos comunes en el inventario:

      - '32 LA ...'            → 32  (pulgadas al inicio con espacio)
      - '32LK540 ...'          → 32  (pulgadas al inicio pegado a letras)
      - 'UN32FH4005G ...'      → 32  (Samsung: UN + pulgadas + modelo)
      - 'KDL-32R425A ...'      → 32  (Sony: KDL- + pulgadas)
      - 'L49S780BTS ...'       → 49  (DAEWOO: L + pulgadas)
      - '49 MASTER G ...'      → 49  (pulgadas al inicio con espacio)
    """
    # 1. Pulgadas al inicio (con espacio o directamente pegadas a letras)
    m = re.match(r"^(\d{2})(?=\s|[A-Za-z])", desc)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 2. Formato Samsung UN: UN{pulgadas}  (ej: UN32FH4005, UN 40 D6500, UN55)
    m = re.search(r"\bUN\s*(\d{2})\s*(?=\d|[A-Za-z])", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 2b. Formato Samsung QN: QN{pulgadas}  (ej: QN55Q60A, QN60Q60A, QN65Q8)
    m = re.search(r"\bQN(\d{2})(?=\d|[A-Za-z])", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 3. Formato Sony: KDL-{pulgadas}  (ej: KDL-32R425A, KDL-39R)
    m = re.search(r"\bKDL[-]?(\d{2})(?=\d|[A-Za-z]|-)", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 4. Formato Daewoo/Toshiba: L{pulgadas}X (ej: L49S780, 43L4700)
    m = re.search(r"\b(?:L|LD)(\d{2})(?=\d|[A-Za-z])", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 5. Formato Daewoo con guión: 43-DAEWO, 49-DAeWO
    m = re.search(r"^(\d{2})-", desc)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 6. Número de 2 dígitos seguido de PULG, comilla de pulgadas, o símbolo "
    m = re.search(r'(\d{2})\s*(?:PULG|"|pulgadas)', desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 7. Formato Sony XBR: XBR-{pulgadas}  (ej: XBR-55W805B)
    m = re.search(r"\bXBR[-]?(\d{2})(?=\d|[A-Za-z])", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 8. Formato JVC LT-: LT-{pulgadas}  (ej: LT-65 KB608)
    m = re.search(r"\bLT[-](\d{2})\b", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    # 9. KDL con espacio en el número: KDL-32 EX30, KDL - 46 R455
    m = re.search(r"\bKDL\s*[-]?\s*(\d{2})\b", desc, re.IGNORECASE)
    if m:
        p = int(m.group(1))
        if 10 <= p <= 100:
            return p

    return None


def parse_leds(desc: str):
    """
    Extrae el número de LEDs por barra desde la descripción.
    Detecta patrones como: 6L, 7LED, 8LEDS, 10L
    """
    m = re.search(r"(\d+)\s*[Ll](?:[Ee][Dd][Ss]?)?(?:\s|/|$|-)", desc)
    if m:
        val = int(m.group(1))
        if 1 <= val <= 50:  # rango razonable para leds por barra
            return val
    return None


def find_header_and_offset(ws):
    """
    Busca la fila de encabezado y el desplazamiento de columnas donde empieza CODIGO.
    Retorna (header_row_num, col_offset).
    """
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(v for v in row if v and "CODIGO" in str(v).upper()):
            for j, v in enumerate(row):
                if v and "CODIGO" in str(v).upper():
                    return i, j
    return None, 0


def extract_products_from_sheet(ws, marca_nombre: str) -> tuple[list, list]:
    """
    Extrae todos los productos de una hoja de Excel.
    Retorna (productos_extraidos, filas_omitidas).
    """
    products = []
    skipped = []

    header_row, col_offset = find_header_and_offset(ws)
    if not header_row:
        skipped.append({"motivo": f"No se encontró encabezado CODIGO en la hoja '{marca_nombre}'"})
        return products, skipped

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_num <= header_row:
            continue

        actual = list(row[col_offset:])
        if len(actual) < 2:
            continue

        raw_codigo = actual[0]
        raw_desc   = actual[1]

        if not raw_codigo or not raw_desc:
            continue

        # Validar que el código sea numérico
        try:
            codigo_viejo = int(raw_codigo)
        except (ValueError, TypeError):
            skipped.append({
                "hoja": marca_nombre,
                "fila": row_num,
                "motivo": f"Código no numérico: '{raw_codigo}'",
            })
            continue

        desc = str(raw_desc).strip()
        if not desc or codigo_viejo <= 0:
            skipped.append({
                "hoja": marca_nombre,
                "fila": row_num,
                "motivo": "Descripción vacía o código <= 0",
            })
            continue

        # Columnas: COD | DESC | STOCK_J | STOCK_B | PRECIO_V_J | PRECIO_C_J | BARRAS | ...
        stock_j  = safe_num(actual[2] if len(actual) > 2 else None, 0)
        stock_b  = safe_num(actual[3] if len(actual) > 3 else None, 0)
        precio_v = safe_num(actual[4] if len(actual) > 4 else None, 0)
        precio_c = safe_num(actual[5] if len(actual) > 5 else None, 0)
        barras_j = safe_num(actual[6] if len(actual) > 6 else None)

        # Normalizar stocks
        stock_j = max(0, int(stock_j)) if (STOCKS_NEGATIVOS_A_CERO and stock_j < 0) else int(stock_j or 0)
        stock_b = max(0, int(stock_b)) if (STOCKS_NEGATIVOS_A_CERO and stock_b < 0) else int(stock_b or 0)

        # Extraer datos técnicos de la descripción
        pulgadas   = parse_pulgadas(desc)
        leds       = parse_leds(desc)
        barras_val = int(barras_j) if barras_j and barras_j > 0 else None

        products.append({
            "marca":                marca_nombre,
            "codigo":               str(codigo_viejo),
            "nombre":               desc,
            "descripcion":          None,
            "modelo_tv":            None,
            "pulgadas_tv":          pulgadas,
            "leds_por_barra":       leds,
            "caracteristicas_barra": None,
            "unidad":               "juego",
            "empaque":              "barra" if barras_val else None,
            "unidades_por_empaque": barras_val,
            "precio_compra":        round(precio_c, 2) if precio_c else 0.0,
            "precio_venta":         round(precio_v, 2) if precio_v else 0.0,
            "stock_actual":         stock_j,
            "stock_barras":         stock_b,
            "stock_minimo":         STOCK_MINIMO_DEFAULT,
            "stock_ideal":          STOCK_IDEAL_DEFAULT,
            "tiempo_reposicion_dias": TIEMPO_REPOSICION_DEFAULT,
            "activo":               True,
        })

    return products, skipped


def main():
    parser = argparse.ArgumentParser(
        description="Extrae productos del Excel de inventario y genera JSON para el seeder Laravel."
    )
    parser.add_argument(
        "--excel",
        default="INVENTARIO 2026 1.xlsx",
        help="Ruta al archivo Excel (default: 'INVENTARIO 2026 1.xlsx')",
    )
    parser.add_argument(
        "--output",
        default="database/seeders/data/productos_importacion.json",
        help="Ruta de salida del JSON (default: database/seeders/data/productos_importacion.json)",
    )
    args = parser.parse_args()

    excel_path = args.excel
    output_path = args.output

    if not os.path.exists(excel_path):
        print(f"ERROR: No se encontró el archivo Excel: '{excel_path}'")
        sys.exit(1)

    print(f"Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_products = []
    all_skipped  = []
    stats_por_marca = {}

    for marca_nombre, sheet_name in MARCA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠  Hoja '{sheet_name}' no encontrada, omitiendo...")
            continue

        ws = wb[sheet_name]
        products, skipped = extract_products_from_sheet(ws, marca_nombre)
        all_products.extend(products)
        all_skipped.extend(skipped)
        stats_por_marca[marca_nombre] = len(products)

    # Verificar códigos duplicados
    codigos = [p["codigo"] for p in all_products]
    duplicados = {c for c in codigos if codigos.count(c) > 1}

    # Crear directorio de salida si no existe
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Guardar JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_products, f, ensure_ascii=False, indent=2)

    # Reporte
    print("\n" + "═" * 55)
    print("  RESULTADO DE EXTRACCIÓN")
    print("═" * 55)
    print(f"{'Marca':<22} {'Productos':>10}")
    print("─" * 55)
    for marca, count in stats_por_marca.items():
        print(f"  {marca:<20} {count:>10}")
    print("─" * 55)
    print(f"  {'TOTAL':<20} {len(all_products):>10}")
    print(f"  {'Omitidos':<20} {len(all_skipped):>10}")
    print(f"  {'Duplicados':<20} {len(duplicados):>10}")

    if duplicados:
        print(f"\n⚠ CÓDIGOS DUPLICADOS DETECTADOS:")
        for d in sorted(duplicados):
            print(f"   - {d}")

    if all_skipped:
        print(f"\n⚠ FILAS OMITIDAS:")
        for s in all_skipped[:10]:
            print(f"   - {s}")
        if len(all_skipped) > 10:
            print(f"   ... y {len(all_skipped) - 10} más.")

    print(f"\n✓ JSON generado en: {output_path}")
    print("═" * 55)
    print(f"\nPara importar al sistema Laravel, ejecuta:")
    print(f"  php artisan db:seed --class=ImportarProductosExcelSeeder")
    print()


if __name__ == "__main__":
    main()
